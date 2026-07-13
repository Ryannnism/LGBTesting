#!/usr/bin/env python3
"""Import Sharon initialization workbook into LGB API (billing parties, customers, division flags)."""

from __future__ import annotations

import argparse
import json
import os
import sys
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

API_BASE = os.environ.get("LGB_API_BASE", "http://localhost:5003").rstrip("/")
PASSWORD = os.environ.get("LGB_DEV_PASSWORD", "password123")
DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parents[1]
    / "docs"
    / "templates"
    / "LGB_Customer_Initialization_Template.xlsx"
)
ADD_ON_UNIT_PRICE = 120

DIVISION_NAME_TO_CODE = {
    "lgb": "LGB",
    "lgb ins": "LGB_INS",
    "exitra": "EXITRA",
    "bellworth & london": "BELLWORTH",
    "uer": "UER",
    "parkwood": "PARKWOOD",
    "swm": "SWM",
    "dlal": "DLAL",
    "dlcm": "DLCM",
    "dlcm & kk": "DLCM_KK",
    "dlcm & sean": "DLCM_SEAN",
    "hold on": "HOLD_ON",
    "nominees": "NOMINEES",
    "tcb": "TCB",
    "kk": "KK",
    "sean": "SEAN",
    "ishak": "ISHAK",
    "janice": "JANICE",
}
for code in list(DIVISION_NAME_TO_CODE.values()):
    DIVISION_NAME_TO_CODE[code.lower()] = code

MOA_LABEL_TO_CODE = {
    "division default": "",
    "moa — no loa": "MOA_NO_LOA",
    "moa - no loa": "MOA_NO_LOA",
    "moa — with loa": "MOA_WITH_LOA",
    "moa - with loa": "MOA_WITH_LOA",
    "moa — swm group": "MOA_SWM",
    "moa - swm group": "MOA_SWM",
    "moa_no_loa": "MOA_NO_LOA",
    "moa_with_loa": "MOA_WITH_LOA",
    "moa_swm": "MOA_SWM",
}

SHEET_ALIASES = {
    "billing": ["Who we invoice", "Billing_Parties"],
    "divisions": ["Divisions", "Ref_Divisions"],
    "moa_setup": ["Company MOA setup"],
    "customers": ["Companies", "Customers"],
    "packages": ["Contracts", "Customer_Packages"],
    "signatories": ["Client contacts", "Customer_Signatories"],
}

BILLING_CATEGORY_MAP = {
    "invoice only": "InvoiceBy",
    "charge only": "ChargeTo",
    "invoice and charge": "Both",
    "invoiceby": "InvoiceBy",
    "chargeto": "ChargeTo",
    "both": "Both",
}


def yn(value: Any) -> bool:
    return str(value or "").strip().upper() in ("Y", "YES", "TRUE", "1")


def cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def row_get(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row:
            val = row[key]
            if val is not None and str(val).strip() != "":
                return val
    return None


def is_skip_row(record: dict[str, Any]) -> bool:
    notes = cell_str(row_get(record, "Notes"))
    if notes.lower().startswith("example"):
        return True
    for key in ("Company name", "CompanyName", "Division", "Name on invoice / account"):
        val = cell_str(row_get(record, key))
        if val.upper().startswith("NOTE"):
            return True
    company = cell_str(row_get(record, "Company name", "CompanyName"))
    if "column guide" in company.lower() or company.endswith("→"):
        return True
    if "Division" in record and not cell_str(record.get("Division")):
        return True
    if cell_str(row_get(record, "Notes")).lower().startswith("confirm managers"):
        return True
    return False


def resolve_sheet_name(wb_path: Path, aliases: list[str]) -> str:
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    names = set(wb.sheetnames)
    wb.close()
    for name in aliases:
        if name in names:
            return name
    raise KeyError(f"None of {aliases!r} found in {wb_path}")


def read_sheet_rows(wb_path: Path, aliases: list[str]) -> list[dict[str, Any]]:
    sheet = resolve_sheet_name(wb_path, aliases)
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []
    headers = [cell_str(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        if is_skip_row(record):
            continue
        out.append(record)
    wb.close()
    return out


def read_moa_setup_by_company(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        company = cell_str(row_get(row, "Company name"))
        if not company:
            continue
        out[company.lower()] = row
    return out


def collect_workflow_names(*values: Any) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for value in values:
        for part in split_names(value):
            key = part.lower()
            if key not in seen:
                seen.add(key)
                names.append(part)
    return names


def resolve_division_code(value: Any) -> str:
    raw = cell_str(value)
    if not raw:
        return ""
    return DIVISION_NAME_TO_CODE.get(raw.lower(), raw.upper())


def resolve_moa_workflow_code(value: Any) -> str | None:
    raw = cell_str(value)
    if not raw:
        return None
    mapped = MOA_LABEL_TO_CODE.get(raw.lower())
    if mapped is not None:
        return mapped or None
    upper = raw.upper()
    if upper in ("MOA_NO_LOA", "MOA_WITH_LOA", "MOA_SWM"):
        return upper
    return None


def resolve_billing_category(value: Any) -> str:
    raw = cell_str(value)
    if not raw:
        return "Both"
    return BILLING_CATEGORY_MAP.get(raw.lower(), raw)


def split_names(value: Any) -> list[str]:
    text = cell_str(value)
    if not text:
        return []
    parts: list[str] = []
    for chunk in text.replace(";", ",").split(","):
        name = chunk.strip()
        if name:
            parts.append(name)
    return parts


def split_emails(value: Any) -> list[str]:
    return [e.strip().lower() for e in split_names(value) if "@" in e]


def build_pricing_json(row: dict[str, Any]) -> str:
    validity = cell_str(row_get(row, "Contract length", "Validity")) or "1 Year"
    base = row_get(row, "Total fee (RM)", "BasePackagePrice_MYR", "ContractValue_MYR") or 0
    try:
        base_price = float(base)
    except (TypeError, ValueError):
        base_price = 0.0
    add_on_lines = []
    for i in range(1, 4):
        name = cell_str(row_get(row, f"Extra service {i}", f"AddOn{i}_Name"))
        qty_raw = row_get(row, f"Qty {i}", f"AddOn{i}_Qty")
        if not name:
            continue
        try:
            qty = int(float(qty_raw or 0))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        add_on_lines.append({"name": name, "qty": qty, "unitPrice": ADD_ON_UNIT_PRICE})
    return json.dumps(
        {"validity": validity, "basePackagePrice": base_price, "addOnLines": add_on_lines}
    )


def request_json(
    method: str,
    path: str,
    body: dict | list | None = None,
    token: str = "",
) -> tuple[int, Any]:
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    data = None if body is None else json.dumps(body).encode()
    req = urllib.request.Request(f"{API_BASE}{path}", data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read().decode()
            return resp.status, json.loads(raw) if raw else None
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode()
        try:
            payload = json.loads(raw) if raw else {"message": exc.reason}
        except json.JSONDecodeError:
            payload = {"message": raw or exc.reason}
        return exc.code, payload


def login(email: str) -> str:
    status, data = request_json("POST", "/api/auth/login", {"email": email, "password": PASSWORD})
    if status != 200 or not data or not data.get("token"):
        raise RuntimeError(f"Login failed for {email}: {data}")
    return data["token"]


def load_users_by_email(token: str) -> dict[str, int]:
    status, users = request_json("GET", "/api/users", token=token)
    if status != 200:
        raise RuntimeError(f"GET users failed: {users}")
    return {
        cell_str(u["email"]).lower(): int(u["userId"])
        for u in users or []
        if u.get("email")
    }


def import_billing_parties(token: str, rows: list[dict[str, Any]]) -> dict[str, int]:
    status, existing = request_json("GET", "/api/billingparties", token=token)
    if status != 200:
        raise RuntimeError(f"GET billing parties failed: {existing}")
    by_name = {cell_str(p["name"]).lower(): p["id"] for p in existing or []}
    for row in rows:
        name = cell_str(row_get(row, "Name on invoice / account", "BillingPartyName"))
        if not name:
            continue
        if name.lower() in by_name:
            continue
        body = {
            "name": name,
            "category": resolve_billing_category(row_get(row, "Used for", "Category")),
            "isActive": yn(row_get(row, "Active") or "Y"),
            "sortOrder": 0,
        }
        status, data = request_json("POST", "/api/billingparties", body, token=token)
        if status not in (200, 201):
            raise RuntimeError(f"Create billing party {name!r} failed: {data}")
        by_name[name.lower()] = data["id"]
    return by_name


def reload_billing_party_ids(token: str) -> dict[str, int]:
    status, existing = request_json("GET", "/api/billingparties", token=token)
    if status != 200:
        raise RuntimeError(f"GET billing parties failed: {existing}")
    return {cell_str(p["name"]).lower(): p["id"] for p in existing or []}


def import_division_config(token: str, rows: list[dict[str, Any]], users_by_email: dict[str, int]) -> int:
    status, groups = request_json("GET", "/api/divisiongroups", token=token)
    if status != 200:
        raise RuntimeError(f"GET division groups failed: {groups}")
    by_name = {cell_str(g["name"]).lower(): g for g in groups or []}
    updated = 0
    for row in rows:
        div_name = cell_str(row_get(row, "Division", "DivisionName"))
        if not div_name:
            continue
        group = by_name.get(div_name.lower())
        if not group:
            print(f"  ! division not found: {div_name!r}")
            continue
        moa_code = resolve_moa_workflow_code(row_get(row, "Default MOA workflow", "MoaWorkflowTemplate"))
        recommenders = []
        for i in range(1, 5):
            name = cell_str(
                row_get(
                    row,
                    f"MOI recommender {i} name",
                    f"MOI manager {i}",
                    f"Recommender{i}",
                )
            )
            email = cell_str(row_get(row, f"MOI recommender {i} email"))
            if not name:
                continue
            user_id = users_by_email.get(email.lower()) if email else None
            recommenders.append({"id": 0, "displayName": name, "userId": user_id})
        body = {
            "id": group["id"],
            "code": group["code"],
            "name": group["name"],
            "moaWorkflowTemplateCode": moa_code or group.get("moaWorkflowTemplateCode") or "MOA_NO_LOA",
            "defaultMoiFormTemplateCode": group.get("defaultMoiFormTemplateCode"),
            "defaultMoaFormTemplateCode": group.get("defaultMoaFormTemplateCode"),
            "isActive": group.get("isActive", True),
            "recommenders": recommenders or group.get("recommenders") or [],
        }
        status, _ = request_json("PUT", f"/api/divisiongroups/{group['id']}", body, token=token)
        if status not in (200, 204):
            raise RuntimeError(f"Update division {div_name!r} failed: {_}")
        updated += 1
        print(f"  ~ division {div_name} ({len(body['recommenders'])} recommenders)")
    return updated


def import_customers(
    token: str,
    customer_rows: list[dict[str, Any]],
    package_rows: list[dict[str, Any]],
    signatory_rows: list[dict[str, Any]],
    party_ids: dict[str, int],
    moa_setup_by_company: dict[str, dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    packages_by_company: dict[str, list[dict[str, Any]]] = {}
    for row in package_rows:
        company = cell_str(row_get(row, "Company name", "CompanyName"))
        if company:
            packages_by_company.setdefault(company, []).append(row)

    signatories_by_company: dict[str, list[dict[str, Any]]] = {}
    for row in signatory_rows:
        company = cell_str(row_get(row, "Company name", "CompanyName"))
        if company:
            signatories_by_company.setdefault(company, []).append(row)

    created_index: dict[str, dict[str, Any]] = {}
    for row in customer_rows:
        company = cell_str(row_get(row, "Company name", "CompanyName"))
        if not company:
            continue
        invoice_name = cell_str(row_get(row, "Invoice to", "InvoiceByParty"))
        charge_name = cell_str(row_get(row, "Bill to", "ChargeToParty"))
        invoice_ids = [party_ids[invoice_name.lower()]] if invoice_name.lower() in party_ids else []
        charge_ids = [party_ids[charge_name.lower()]] if charge_name.lower() in party_ids else []
        if not invoice_ids or not charge_ids:
            raise RuntimeError(
                f"Customer {company!r}: invoice/charge party not found "
                f"({invoice_name!r}, {charge_name!r}). Fill Who we invoice first."
            )

        setup = moa_setup_by_company.get(company.lower(), {})
        loa_names = collect_workflow_names(
            row_get(row, "LOA holder names (if Y)", "LOAHolderNames"),
            row_get(setup, "LOA holder 1 name (auto)", "LOA holder 1 name"),
            row_get(setup, "LOA holder 2 name"),
            row_get(setup, "Board member 1 name"),
            row_get(setup, "Board member 2 name"),
        )
        moa_code = resolve_moa_workflow_code(
            row_get(setup, "MOA workflow", "MOA workflow (from MOA setup)")
            or row_get(row, "MOA workflow", "MOA workflow (from MOA setup)", "MoaWorkflowOverride")
        )

        holders = []
        for sig in signatories_by_company.get(company, []):
            name = cell_str(row_get(sig, "Person name", "FullName"))
            if not name:
                continue
            holders.append(
                {
                    "id": 0,
                    "name": name,
                    "email": cell_str(row_get(sig, "Email")),
                    "phone": cell_str(row_get(sig, "Phone", "Mobile")),
                    "moi": yn(row_get(sig, "Fills in MOI?", "MOI_Issuer")),
                    "moiApproval": yn(row_get(sig, "Approves MOI (client)?", "MOI_Approver")),
                    "moa": yn(row_get(sig, "Signs MOA?", "MOA_Approver")),
                }
            )

        packages = []
        for pkg in packages_by_company.get(company, []):
            packages.append(
                {
                    "packageName": cell_str(row_get(pkg, "Package", "PackageName")),
                    "packageValue": cell_str(row_get(pkg, "Total fee (RM)", "ContractValue_MYR")),
                    "packageDetail": cell_str(row_get(pkg, "Notes", "PackageDetail")) or None,
                    "validity": cell_str(row_get(pkg, "Contract length", "Validity")) or "1 Year",
                    "purchasedDate": cell_str(row_get(pkg, "Start date", "PurchasedDate")),
                    "pricingJson": build_pricing_json(pkg),
                    "status": cell_str(row_get(pkg, "PackageStatus")) or "Active",
                }
            )

        body: dict[str, Any] = {
            "companyName": company,
            "contactName": cell_str(row_get(row, "Main contact name", "PrimaryContactName")),
            "email": cell_str(row_get(row, "Main contact email", "PrimaryContactEmail")),
            "mobile": cell_str(row_get(row, "Main contact phone", "PrimaryContactMobile")),
            "cosec": yn(row_get(row, "COSEC client?", "CosecCustomer")),
            "divisionGroupCode": resolve_division_code(row_get(row, "Division", "DivisionCode")),
            "hasLoa": yn(row_get(row, "Has Limit of Authority (LOA)?", "HasLOA")),
            "loaHolders": loa_names,
            "invoiceBy": invoice_name,
            "chargeTo": charge_name,
            "invoiceByPartyIds": invoice_ids,
            "chargeToPartyIds": charge_ids,
            "dateCreated": cell_str(row_get(row, "DateCreated")) or None,
            "moaWorkflowTemplateCode": moa_code,
            "packages": packages,
            "accountHolders": holders,
        }
        status, data = request_json("POST", "/api/customers", body, token=token)
        if status not in (200, 201):
            raise RuntimeError(f"Create customer {company!r} failed ({status}): {data}")
        pkg_index = {
            cell_str(p.get("packageName", "")).lower(): int(p["id"])
            for p in data.get("packages") or []
            if p.get("id")
        }
        created_index[company.lower()] = {
            "id": data.get("id"),
            "company": company,
            "packages": pkg_index,
        }
        moa_label = moa_code or "division default"
        print(f"  + customer {company} (id={data.get('id')}, MOA={moa_label})")

    return created_index


def import_division_flags(token: str, customer_rows: list[dict[str, Any]]) -> Any:
    rows = []
    for row in customer_rows:
        company = cell_str(row_get(row, "Company name", "CompanyName"))
        division = resolve_division_code(row_get(row, "Division", "DivisionCode"))
        if not company or not division:
            continue
        rows.append(
            {
                "company": company,
                "divisionGroup": division,
                "hasLoa": yn(row_get(row, "Has Limit of Authority (LOA)?", "HasLOA")),
            }
        )
    if not rows:
        return None
    status, data = request_json("POST", "/api/divisiongroups/import", rows, token=token)
    if status != 200:
        raise RuntimeError(f"Division import failed: {data}")
    return data


def import_contract_staff_tags(
    token: str,
    package_rows: list[dict[str, Any]],
    customer_index: dict[str, dict[str, Any]],
    users_by_email: dict[str, int],
) -> int:
    tagged = 0
    for row in package_rows:
        company = cell_str(row_get(row, "Company name", "CompanyName"))
        package_name = cell_str(row_get(row, "Package", "PackageName"))
        emails = split_emails(
            row_get(row, "Additional LGB staff to tag (emails)", "AdditionalStaffEmails")
        )
        if not company or not package_name or not emails:
            continue
        customer = customer_index.get(company.lower())
        if not customer:
            print(f"  ! contract tag skipped — customer not found: {company!r}")
            continue
        package_id = customer["packages"].get(package_name.lower())
        if not package_id:
            print(f"  ! contract tag skipped — package not found: {company!r} / {package_name!r}")
            continue
        status, jobs = request_json(
            "GET",
            f"/api/jobrequests?customerPackageId={package_id}&includeCompleted=true",
            token=token,
        )
        if status != 200:
            print(f"  ! jobs lookup failed for package {package_id}: {jobs}")
            continue
        for job in jobs or []:
            if job.get("taskType") != "Service":
                continue
            job_id = job.get("id") or job.get("jobRequestId")
            if not job_id:
                continue
            for email in emails:
                user_id = users_by_email.get(email)
                if not user_id:
                    print(f"  ! unknown staff email (skipped): {email}")
                    continue
                assign_body = {"userId": user_id, "unitNumber": 1}
                a_status, a_data = request_json(
                    "POST",
                    f"/api/jobrequests/{job_id}/assign",
                    assign_body,
                    token=token,
                )
                if a_status == 200:
                    tagged += 1
                else:
                    msg = a_data.get("message") if isinstance(a_data, dict) else a_data
                    print(f"  ! assign {email} to job {job_id}: {msg}")
    return tagged


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--email", default=os.environ.get("LGB_ADMIN_EMAIL", "sharon@lgb.test"))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        print("Run: python3 scripts/build_sharon_init_template.py", file=sys.stderr)
        sys.exit(1)

    billing = read_sheet_rows(args.workbook, SHEET_ALIASES["billing"])
    divisions = read_sheet_rows(args.workbook, SHEET_ALIASES["divisions"])
    moa_setup_rows = read_sheet_rows(args.workbook, SHEET_ALIASES["moa_setup"])
    moa_setup_by_company = read_moa_setup_by_company(moa_setup_rows)
    customers = read_sheet_rows(args.workbook, SHEET_ALIASES["customers"])
    packages = read_sheet_rows(args.workbook, SHEET_ALIASES["packages"])
    signatories = read_sheet_rows(args.workbook, SHEET_ALIASES["signatories"])

    print(f"Workbook: {args.workbook}")
    print(
        f"  billing={len(billing)} divisions={len(divisions)} "
        f"moa_setup={len(moa_setup_rows)} companies={len(customers)} "
        f"contracts={len(packages)} contacts={len(signatories)}"
    )
    if args.dry_run:
        print("Dry run — no API calls.")
        return

    token = login(args.email)
    print(f"Logged in as {args.email}")
    users_by_email = load_users_by_email(token)

    print("Updating divisions (MOI managers + default MOA workflow)…")
    div_count = import_division_config(token, divisions, users_by_email)
    print(f"Updated {div_count} division(s).")

    print("Importing billing parties…")
    import_billing_parties(token, billing)
    party_ids = reload_billing_party_ids(token)

    print("Importing customers…")
    customer_index = import_customers(
        token, customers, packages, signatories, party_ids, moa_setup_by_company
    )
    print(f"Created {len(customer_index)} customer(s).")

    print("Syncing division flags…")
    result = import_division_flags(token, customers)
    if result:
        print(f"  division import: updated={result.get('updated')} unmatched={result.get('unmatched')}")

    print("Tagging additional contract staff…")
    tag_count = import_contract_staff_tags(token, packages, customer_index, users_by_email)
    print(f"Tagged {tag_count} job assignment(s).")

    print("Done.")


if __name__ == "__main__":
    main()
