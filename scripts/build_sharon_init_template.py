#!/usr/bin/env python3
"""Generate Sharon-friendly customer data workbook with per-company MOA role formulas."""

from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Install openpyxl: python3 -m pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "templates" / "LGB_Customer_Initialization_Template.xlsx"
CSV_DIR = ROOT / "docs" / "templates" / "csv"
OPEN_ME = ROOT / "docs" / "templates" / "OPEN_IN_EXCEL.txt"

DIVISIONS = [
    ("LGB", "LGB", "MOA — No LOA", ["Tai", "Sam"]),
    ("LGB_INS", "LGB INS", "MOA — No LOA", ["Steven", "Tiew Siong Yi"]),
    ("EXITRA", "Exitra", "MOA — No LOA", ["Kevin Teoh", "Tay", "Jonas"]),
    ("BELLWORTH", "Bellworth & London", "MOA — No LOA", ["Kevin Kuok", "Jaslyn", "Wong Wai Leng", "Evelyn"]),
    ("UER", "UER", "MOA — No LOA", ["Sam Lau"]),
    ("PARKWOOD", "Parkwood", "MOA — No LOA", ["Danny Ng", "Casper Wong"]),
    ("SWM", "SWM", "MOA — SWM Group", ["Shirley", "Bin", "Yvonne", "Thomas"]),
    ("DLAL", "DLAL", "MOA — No LOA", []),
    ("DLCM", "DLCM", "MOA — With LOA", []),
    ("DLCM_KK", "DLCM & KK", "MOA — With LOA", []),
    ("DLCM_SEAN", "DLCM & SEAN", "MOA — With LOA", []),
    ("HOLD_ON", "HOLD ON", "MOA — No LOA", []),
    ("NOMINEES", "Nominees", "MOA — No LOA", []),
    ("TCB", "TCB", "MOA — No LOA", []),
    ("KK", "KK", "MOA — With LOA", []),
    ("SEAN", "SEAN", "MOA — With LOA", []),
    ("ISHAK", "ISHAK", "MOA — No LOA", []),
    ("JANICE", "JANICE", "MOA — No LOA", []),
]

MOA_WORKFLOW_OPTIONS = [
    "MOA — No LOA",
    "MOA — With LOA",
    "MOA — SWM Group",
]

# Standardised steps (mirrors WorkflowConfigSeeder)
MOA_WORKFLOW_STEPS = {
    "MOA — No LOA": [
        (1, "Senior Manager, Company Secretarial", "LGB internal (fixed)"),
        (2, "Project initiator", "Per company — client contact"),
        (3, "Head of Finance / CFO", "Per company — when finance-related"),
        (4, "CEO / COO / GM", "Per company — when finance / share movement"),
        (5, "Ms Teh", "LGB internal (bank signatory matters)"),
        (6, "Full Board", "Per company — board member names"),
        (7, "DLCM", "LGB / external (fixed)"),
    ],
    "MOA — With LOA": [
        (1, "Senior Manager, Company Secretarial", "LGB internal (fixed)"),
        (2, "Project initiator", "Per company — client contact"),
        (3, "Head of Finance / CFO", "Per company — when finance-related"),
        (4, "Ms Teh", "LGB internal (bank signatory matters)"),
        (5, "DLCM — all bank matters", "LGB / external (fixed)"),
    ],
    "MOA — SWM Group": [
        (1, "Senior Manager, Company Secretarial", "LGB internal (fixed)"),
        (2, "Project initiator", "Per company — client contact"),
        (3, "Head of Finance / CFO", "Per company — when finance-related"),
        (4, "LOA / BOD / EXCO holders", "Per company — LOA holder names"),
        (5, "Regulator & Compliance", "LGB internal (Shirley Nicholas)"),
        (6, "CEO / COO / Janice Lim", "LGB internal (fixed)"),
        (7, "Ms Teh", "LGB internal (bank signatory matters)"),
    ],
}

# Role row label -> values per workflow column for formula matrix
ROLE_MATRIX_WORKFLOWS = MOA_WORKFLOW_OPTIONS
ROLE_MATRIX_ROWS = [
    ("Project initiator", ["Fill", "Fill", "Fill"]),
    ("MOA signers (client)", ["Auto", "Auto", "Auto"]),
    ("CFO / Head of Finance", ["If finance", "If finance", "If finance"]),
    ("CEO / COO / GM", ["If finance", "—", "—"]),
    ("Board member 1", ["Fill", "—", "—"]),
    ("Board member 2", ["Fill", "—", "—"]),
    ("LOA holder 1", ["—", "—", "Fill"]),
    ("LOA holder 2", ["—", "—", "Fill"]),
    ("Senior Manager CoSec", ["LGB fixed", "LGB fixed", "LGB fixed"]),
    ("Division MOI managers", ["Auto", "Auto", "Auto"]),
    ("Ms Teh (bank)", ["If bank", "If bank", "If bank"]),
    ("DLCM", ["LGB fixed", "LGB fixed", "—"]),
]

PACKAGES = [
    ("Dormant", 2360),
    ("Basic Package", 3200),
    ("Professional Package", 4250),
    ("Enterprise Package", 5930),
    ("Enterprise Plus", 6980),
    ("Add-ons only", 0),
]

ADD_ONS = [
    "Overseas Support Service",
    "Local Support Service",
    "Attend Board Meeting",
    "Prepare board meeting Minutes",
    "Lodgement fee to MBSR Audited A/C",
    "Lodgement fee for MBRS annual return",
]

HEADER_FILL = PatternFill("solid", fgColor="FFF2CC")
REQUIRED_FILL = PatternFill("solid", fgColor="FFE699")
REF_FILL = PatternFill("solid", fgColor="E2EFDA")
LOOKUP_FILL = PatternFill("solid", fgColor="D9E1F2")
HINT_FILL = PatternFill("solid", fgColor="FCE4D6")
WRAP = Alignment(wrap_text=True, vertical="top")
YN = '"Y,N"'
MOA_DV = ",".join(f'"{o}"' for o in MOA_WORKFLOW_OPTIONS)
DIV_DATA_START = 3
MOA_SETUP_START = 3
MATRIX_DATA_START = 2


def style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = WRAP


def mark_required(ws, cols: list[int], row: int = 1) -> None:
    for col in cols:
        ws.cell(row=row, column=col).fill = REQUIRED_FILL


def set_col_widths(ws, widths: list[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_list_validation(ws, cell_range: str, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please pick from the dropdown."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def matrix_hint(role_label: str, row: int, workflow_col: str = "B") -> str:
    """Row 2 hint: what Sharon should do for this column based on workflow pick."""
    role_row = MATRIX_DATA_START + next(
        i for i, (label, _) in enumerate(ROLE_MATRIX_ROWS) if label == role_label
    )
    col_match = f"MATCH(${workflow_col}{row},'MOA role matrix'!$B$1:$D$1,0)"
    cell = f"INDEX('MOA role matrix'!$B${MATRIX_DATA_START}:$D${MATRIX_DATA_START + len(ROLE_MATRIX_ROWS) - 1},{role_row - MATRIX_DATA_START + 1},{col_match})"
    return f'=IF(${workflow_col}{row}="","",{cell})'


def contact_pick(company_row: int, flag_col: str, value_col: str, which: int = 1) -> str:
    """Pick nth matching client contact (Excel 365 FILTER)."""
    filt = (
        f"FILTER('Client contacts'!${value_col}$2:${value_col}$500,"
        f"('Client contacts'!$A$2:$A$500=$A{company_row})*"
        f"('Client contacts'!${flag_col}$2:${flag_col}$500=\"Y\"),\"\")"
    )
    if which == 1:
        return f'=IF($A{company_row}="","",IFERROR(INDEX({filt},1),""))'
    return f'=IF($A{company_row}="","",IFERROR(INDEX({filt},{which}),""))'


def roles_needed_formula(row: int) -> str:
    parts: list[str] = []
    for i, (role_name, _) in enumerate(ROLE_MATRIX_ROWS):
        idx = i + 1
        col_match = f"MATCH($B{row},'MOA role matrix'!$B$1:$D$1,0)"
        val = (
            f"INDEX('MOA role matrix'!$B${MATRIX_DATA_START}:$D$"
            f"{MATRIX_DATA_START + len(ROLE_MATRIX_ROWS) - 1},{idx},{col_match})"
        )
        parts.append(f'IF({val}="Fill","{role_name}, ","")')
        parts.append(f'IF({val}="If finance","{role_name} (if finance), ","")')
        parts.append(f'IF({val}="Auto","{role_name} (auto), ","")')
    joined = "".join(parts)
    return f'=IF($B{row}="","",TRIM({joined}))'


def division_moi_lookup_formula(row: int) -> str:
    match = f"MATCH($E{row},Divisions!$A${DIV_DATA_START}:$A$50,0)"
    cols = ("C", "E", "G", "I")
    parts = [f'IFERROR(INDEX(Divisions!${c}${DIV_DATA_START}:${c}$50,{match}),"")' for c in cols]
    return (
        f'=IF($E{row}="","",TRIM({parts[0]}'
        f'&IF({parts[1]}<>"",", "&{parts[1]},"")'
        f'&IF({parts[2]}<>"",", "&{parts[2]},"")'
        f'&IF({parts[3]}<>"",", "&{parts[3]},"")))'
    )


def company_moa_workflow_lookup(row: int) -> str:
    return (
        f"=IFERROR(INDEX('Company MOA setup'!$B${MOA_SETUP_START}:$B$500,"
        f"MATCH($A{row},'Company MOA setup'!$A${MOA_SETUP_START}:$A$500,0)),"
        f"\"Division default\")"
    )


def sheet_start_here(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Start here"
    lines = [
        "LGB — customer data collection (for Sharon)",
        "",
        "Open in Excel or Numbers — not in a code editor. Formulas need Excel 365 or Excel 2021+.",
        "",
        "Fill in this order:",
        "  1. Who we invoice",
        "  2. Divisions — MOI managers + emails",
        "  3. Companies — basics (division, LOA, billing)",
        "  4. Client contacts — who fills MOI, signs MOA, starts requests",
        "  5. Company MOA setup — pick workflow per company; fill name/email blanks",
        "  6. Contracts — packages + extra LGB staff only",
        "",
        "MOA workflows (read MOA workflow steps + MOA role matrix tabs):",
        "  • Each company can use a different MOA workflow.",
        "  • Row 2 on Company MOA setup shows what each column needs (Fill / Auto / —).",
        "  • Blue cells auto-pull from Client contacts or Companies.",
        "  • Yellow cells = Sharon fills in names/emails for that company's approvers.",
        "  • CEO, CFO, board, LOA holders differ per company — only fill columns that apply.",
        "",
        "Delete example rows before import.",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 96
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)


def sheet_moa_workflow_steps(wb: Workbook) -> None:
    ws = wb.create_sheet("MOA workflow steps")
    ws.append(["MOA workflow", "Step", "Approver role", "Who Sharon fills"])
    style_header_row(ws, 1, 4)
    for workflow, steps in MOA_WORKFLOW_STEPS.items():
        for order, role, who in steps:
            ws.append([workflow, order, role, who])
            for cell in ws[ws.max_row]:
                cell.fill = REF_FILL
    ws.freeze_panes = "A2"
    set_col_widths(ws, [22, 6, 36, 36])


def sheet_moa_role_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("MOA role matrix")
    ws.append(["Role slot"] + ROLE_MATRIX_WORKFLOWS)
    style_header_row(ws, 1, 1 + len(ROLE_MATRIX_WORKFLOWS))
    for role, values in ROLE_MATRIX_ROWS:
        ws.append([role, *values])
        for cell in ws[ws.max_row]:
            cell.fill = REF_FILL
    ws.cell(row=MATRIX_DATA_START + len(ROLE_MATRIX_ROWS) + 1, column=1, value=(
        "Legend: Fill = Sharon types name/email | Auto = formula from contacts | "
        "If finance = only when matter is finance-related | LGB fixed = system handles"
    ))
    ws.sheet_state = "hidden"
    set_col_widths(ws, [28, 16, 16, 16])


def sheet_company_moa_setup(wb: Workbook) -> None:
    ws = wb.create_sheet("Company MOA setup")
    headers = [
        "Company name",
        "MOA workflow",
        "Roles in this workflow (auto)",
        "Project initiator name (auto)",
        "Project initiator email (auto)",
        "MOA signer 1 name (auto)",
        "MOA signer 1 email (auto)",
        "MOA signer 2 name (auto)",
        "MOA signer 2 email (auto)",
        "CFO / Finance name",
        "CFO / Finance email",
        "CEO / GM name",
        "CEO / GM email",
        "Board member 1 name",
        "Board member 1 email",
        "Board member 2 name",
        "Board member 2 email",
        "LOA holder 1 name (auto)",
        "LOA holder 1 email",
        "LOA holder 2 name",
        "LOA holder 2 email",
        "LGB prepared-by email",
        "LGB vetted-by email",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    mark_required(ws, [1, 2])

    hint_cols = {
        4: "Project initiator",
        6: "MOA signers (client)",
        10: "CFO / Head of Finance",
        12: "CEO / COO / GM",
        14: "Board member 1",
        16: "Board member 2",
        18: "LOA holder 1",
        20: "LOA holder 2",
    }
    for col, role in hint_cols.items():
        ws.cell(row=2, column=col, value=matrix_hint(role, MOA_SETUP_START))
        ws.cell(row=2, column=col).fill = HINT_FILL
        ws.cell(row=2, column=col).font = Font(italic=True, size=9)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1)
    ws.cell(row=2, column=1, value="Column guide (per workflow) →").fill = HINT_FILL

    example_row = MOA_SETUP_START
    ws.append([
        "Acme Corp",
        "MOA — No LOA",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "Jane CEO",
        "ceo@acme.com",
        "Director One",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "nadia@lgb.test",
        "nita@lgb.test",
        "Example — delete this row",
    ])

    auto_cols = {3, 4, 5, 6, 7, 8, 9, 18}
    fill_cols = {10, 11, 12, 13, 14, 15, 16, 17, 19, 20, 21, 22, 23}
    for row_idx in range(MOA_SETUP_START, 501):
        ws.cell(row=row_idx, column=3).value = roles_needed_formula(row_idx)
        ws.cell(row=row_idx, column=4).value = contact_pick(row_idx, "H", "B", 1)
        ws.cell(row=row_idx, column=5).value = contact_pick(row_idx, "H", "C", 1)
        ws.cell(row=row_idx, column=6).value = contact_pick(row_idx, "G", "B", 1)
        ws.cell(row=row_idx, column=7).value = contact_pick(row_idx, "G", "C", 1)
        ws.cell(row=row_idx, column=8).value = contact_pick(row_idx, "G", "B", 2)
        ws.cell(row=row_idx, column=9).value = contact_pick(row_idx, "G", "C", 2)
        loa_auto = (
            f'=IF($A{row_idx}="","",IFERROR(INDEX(Companies!$I$2:$I$500,'
            f"MATCH($A{row_idx},Companies!$A$2:$A$500,0)),\"\"))"
        )
        ws.cell(row=row_idx, column=18).value = loa_auto
        for col in auto_cols:
            ws.cell(row=row_idx, column=col).fill = LOOKUP_FILL
        for col in fill_cols:
            if row_idx == example_row:
                continue
            ws.cell(row=row_idx, column=col).fill = HEADER_FILL

    ws.freeze_panes = "A4"
    set_col_widths(ws, [20, 20, 34, 22, 26, 20, 26, 20, 26, 18, 24, 18, 24, 18, 24, 18, 24, 20, 24, 18, 24, 22, 22, 24])
    add_list_validation(ws, f"B{MOA_SETUP_START}:B500", MOA_DV)


def sheet_divisions(wb: Workbook) -> None:
    ws = wb.create_sheet("Divisions")
    headers = [
        "Division",
        "Default MOA workflow",
        "MOI recommender 1 name",
        "MOI recommender 1 email",
        "MOI recommender 2 name",
        "MOI recommender 2 email",
        "MOI recommender 3 name",
        "MOI recommender 3 email",
        "MOI recommender 4 name",
        "MOI recommender 4 email",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    mark_required(ws, [2] + list(range(3, 11)))
    for _code, name, moa, recs in DIVISIONS:
        row = [name, moa]
        for i in range(4):
            row.append(recs[i] if i < len(recs) else "")
            row.append("")
        row.append("")
        ws.append(row)
    ws.freeze_panes = "A3"
    set_col_widths(ws, [22, 22, 20, 26, 20, 26, 20, 26, 20, 26, 36])
    add_list_validation(ws, f"B{DIV_DATA_START}:B50", MOA_DV)
    for row_idx in range(DIV_DATA_START, DIV_DATA_START + len(DIVISIONS)):
        ws.cell(row=row_idx, column=1).fill = REF_FILL


def sheet_package_list(wb: Workbook) -> None:
    ws = wb.create_sheet("Package list")
    ws.append(["Package name", "Annual fee (RM)", "Notes"])
    style_header_row(ws, 1, 3)
    for name, price in PACKAGES:
        note = "For non-COSEC clients buying add-ons only" if name == "Add-ons only" else "Standard package"
        ws.append([name, price, note])
    ws.append([])
    ws.append(["Extra service (add-on)", "Unit", ""])
    style_header_row(ws, ws.max_row, 3)
    units = ["Month", "Each", "2hrs/Each", "Per meeting", "Each", "Each"]
    for name, unit in zip(ADD_ONS, units):
        ws.append([name, unit, ""])
    ws.freeze_panes = "A2"
    set_col_widths(ws, [36, 14, 40])
    for row in ws.iter_rows(min_row=2, max_row=len(PACKAGES) + 1):
        for cell in row:
            cell.fill = REF_FILL


def sheet_who_we_invoice(wb: Workbook) -> None:
    ws = wb.create_sheet("Who we invoice")
    headers = ["Name on invoice / account", "Used for", "Active", "Notes"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    mark_required(ws, [1, 2])
    ws.append(["Acme Corp", "Invoice and charge", "Y", "Example — delete this row"])
    ws.freeze_panes = "A2"
    set_col_widths(ws, [32, 22, 8, 30])
    add_list_validation(ws, "C2:C500", YN)
    add_list_validation(ws, "B2:B500", '"Invoice only,Charge only,Invoice and charge"')


def sheet_companies(wb: Workbook) -> None:
    ws = wb.create_sheet("Companies")
    headers = [
        "Company name",
        "Main contact name",
        "Main contact email",
        "Main contact phone",
        "Division",
        "MOA workflow (from MOA setup)",
        "COSEC client?",
        "Has Limit of Authority (LOA)?",
        "LOA holder names (if Y)",
        "Invoice to",
        "Bill to",
        "Division MOI managers (auto)",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    mark_required(ws, [1, 2, 3, 4, 5, 7, 8, 10, 11])
    ws.append([
        "Acme Corp",
        "Sarah Johnson",
        "sarah.j@acme.com",
        "+60 12-345 6789",
        "LGB",
        "",
        "Y",
        "N",
        "",
        "Acme Corp",
        "Acme Corp",
        "",
        "Example — delete this row",
    ])
    for row_idx in range(2, 501):
        ws.cell(row=row_idx, column=6).value = company_moa_workflow_lookup(row_idx)
        ws.cell(row=row_idx, column=6).fill = LOOKUP_FILL
        ws.cell(row=row_idx, column=12).value = division_moi_lookup_formula(row_idx)
        ws.cell(row=row_idx, column=12).fill = LOOKUP_FILL
    ws.freeze_panes = "A2"
    set_col_widths(ws, [22, 18, 26, 16, 22, 28, 12, 10, 28, 20, 20, 28, 24])
    div_names = ",".join(f'"{name}"' for _c, name, _m, _r in DIVISIONS)
    add_list_validation(ws, "E2:E500", div_names)
    add_list_validation(ws, "G2:H500", YN)


def sheet_contracts(wb: Workbook) -> None:
    ws = wb.create_sheet("Contracts")
    headers = [
        "Company name",
        "Package",
        "Contract length",
        "Start date",
        "End date (optional)",
        "Total fee (RM)",
        "Extra service 1",
        "Qty 1",
        "Extra service 2",
        "Qty 2",
        "Extra service 3",
        "Qty 3",
        "Additional LGB staff to tag (emails)",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    mark_required(ws, [1, 2, 3, 4, 6])
    ws.insert_rows(2)
    ws.merge_cells("A2:N2")
    note_cell = ws.cell(
        row=2,
        column=1,
        value=(
            "NOTE: MOA approvers are on Company MOA setup + Client contacts. "
            "Here, only tag extra internal staff emails (e.g. nadia@lgb.test)."
        ),
    )
    note_cell.fill = REF_FILL
    note_cell.alignment = WRAP
    ws.row_dimensions[2].height = 32
    ws.append([
        "Acme Corp",
        "Enterprise Plus",
        "1 Year",
        "2026-01-15",
        "2027-01-14",
        6980,
        "Attend Board Meeting",
        4,
        "Prepare board meeting Minutes",
        2,
        "",
        "",
        "",
        "Example — delete this row",
    ])
    ws.freeze_panes = "A3"
    set_col_widths(ws, [20, 22, 14, 12, 14, 14, 28, 8, 28, 8, 28, 8, 34, 22])
    pkg_names = ",".join(f'"{p[0]}"' for p in PACKAGES)
    add_list_validation(ws, "B3:B500", pkg_names)
    add_list_validation(ws, "C3:C500", '"6 Months,1 Year,2 Years,3 Years"')


def sheet_client_contacts(wb: Workbook) -> None:
    ws = wb.create_sheet("Client contacts")
    headers = [
        "Company name",
        "Person name",
        "Email",
        "Phone",
        "Fills in MOI?",
        "Approves MOI (client)?",
        "Signs MOA?",
        "Usually starts the request?",
        "Notes",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    mark_required(ws, list(range(1, 8)))
    examples = [
        ("Acme Corp", "Dan Ra", "dra@lgb.test", "+60 11-111 1111", "Y", "Y", "Y", "Y", "Example — delete"),
        ("Acme Corp", "Daniel Ra", "dra2@lgb.test", "", "Y", "N", "Y", "N", "Example — delete"),
        ("Acme Corp", "Dra Three", "dra3@lgb.test", "", "N", "Y", "Y", "N", "Example — delete"),
    ]
    for row in examples:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    set_col_widths(ws, [20, 20, 26, 16, 14, 20, 12, 22, 20])
    for col_letter in ("E", "F", "G", "H"):
        add_list_validation(ws, f"{col_letter}2:{col_letter}500", YN)


def export_csv_previews(wb: Workbook) -> None:
    CSV_DIR.mkdir(parents=True, exist_ok=True)
    current = set()
    for ws in wb.worksheets:
        path = CSV_DIR / f"{ws.title}.csv"
        current.add(path.name)
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        print(f"Wrote {path}")
    for stale in CSV_DIR.glob("*.csv"):
        if stale.name not in current:
            stale.unlink()
            print(f"Removed stale {stale}")


def write_open_instructions() -> None:
    OPEN_ME.write_text(
        """LGB customer data workbook (for Sharon)
==========================================

Open in Excel — NOT in Cursor (you will see gibberish).
Use Excel 365 / 2021+ so MOA setup formulas work (FILTER, etc.).

Key tab: Company MOA setup — pick workflow per company, fill approver blanks.

File:
  docs/templates/LGB_Customer_Initialization_Template.xlsx

CSV previews (formulas show as text):
  docs/templates/csv/

Regenerate:
  python3 scripts/build_sharon_init_template.py
""",
        encoding="utf-8",
    )
    print(f"Wrote {OPEN_ME}")


def main() -> None:
    wb = Workbook()
    sheet_start_here(wb)
    sheet_moa_workflow_steps(wb)
    sheet_moa_role_matrix(wb)
    sheet_divisions(wb)
    sheet_who_we_invoice(wb)
    sheet_companies(wb)
    sheet_client_contacts(wb)
    sheet_company_moa_setup(wb)
    sheet_contracts(wb)
    sheet_package_list(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    export_csv_previews(wb)
    wb.save(OUT)
    write_open_instructions()
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
